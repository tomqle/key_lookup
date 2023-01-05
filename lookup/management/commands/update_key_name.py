from django.core.management.base import BaseCommand
from logging import getLogger

from lookup.models import Key, KeyShell, Remote, RemoteShell, EmergencyKey, VehicleApplication

import openpyxl
from datetime import datetime
import pytz

LOG = getLogger(__name__)

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str)

    def handle(self, *args, **options):
        try:
            file_name = options['file']
            if file_name:
                wb = openpyxl.load_workbook(file_name)
                sheet1 = wb[wb.sheetnames[0]]

                print('\nExcel spreadsheet loaded successfully.')

                if self._index_spreadsheet_columns(sheet1):
                    product_data = self._read_product_data_from_excel(sheet1)

                    print(product_data)

                    if product_data:
                        self._bulk_update_key_name(product_data)

                    #timestamp = datetime.now(pytz.timezone('US/Pacific')).strftime('%Y%m%d_%H%M%S')
                    #output_file = 'static/excel/output_file_' + timestamp + '.xlsx'

                    #self._generate_product_data_output_workbook(product_dict, output_file)

                    #print('\nImport results saved to ' + output_file + '.')
        except ValueError as e:
            pass

    def _index_spreadsheet_columns(self, sheet1):
        self.COL_INDEX_ID_1 = -1
        self.COL_INDEX_NAME_1 = -1
        self.COL_INDEX_TYPE_1 = -1

        i = 1
        while True:
            col_name = sheet1.cell(1, i).value
            if col_name == None:
                break
            elif col_name.lower() == 'id':
                self.COL_INDEX_ID_1 = i
            elif col_name.lower() == 'name':
                self.COL_INDEX_NAME_1 = i
            elif col_name.lower() == 'type':
                self.COL_INDEX_TYPE_1 = i

            i += 1
        
        if self.COL_INDEX_ID_1 == -1:
            print('\nID column missing.')
            return False
        elif self.COL_INDEX_NAME_1 == -1:
            print('\nName column missing.')
            return False
        elif self.COL_INDEX_TYPE_1 == -1:
            print('\nType column missing.')
            return False

        print('\nAll columns are present')
        return True

    def _read_product_data_from_excel(self, sheet):
        product_data = []
        for row in range(2, sheet.max_row + 1):
            if isinstance(sheet.cell(row, self.COL_INDEX_ID_1).value, float):
                id1 = int(sheet.cell(row, self.COL_INDEX_ID_1).value)
                print('ID field is float')
            elif isinstance(sheet.cell(row, self.COL_INDEX_ID_1).value, int):
                id1 = sheet.cell(row, self.COL_INDEX_ID_1).value
                print('ID field is int')
            else:
                print('ID field is invalid')
                raise ValueError('ID field value is invalid')

            name = str(sheet.cell(row, self.COL_INDEX_NAME_1).value)
            typename = str(sheet.cell(row, self.COL_INDEX_TYPE_1).value)
            product_data.append({   'id': id1,
                                    'name': name,
                                    'type': typename })

        print('\nProduct data read from excel successfully.')
        print(product_data)

        return product_data
    
    def _bulk_update_key_name(self, product_data):
        products = []
        keys = []
        key_shells = []
        remote_shells = []
        emergency_keys = []

        for p in product_data:
            product.name = p['name']

            if p['type'].lower() == 'key' or p['type'].lower() == 'remote':
                product = Key.objects.get(id=p['id'])
                keys.append(product)
            elif p['type'].lower() == 'key shell':
                product = KeyShell.objects.get(id=p['id'])
                key_shells.append(product)
            elif p['type'].lower() == 'remote shell':
                product = RemoteShell.objects.get(id=p['id'])
                remote_shells.append(product)
            elif p['type'].lower() == 'emergency key':
                product = EmergencyKey.objects.get(id=p['id'])
                emergency_keys.append(product)
            else:
               raise ValueError('Type field value is invalid')

        if keys:
            Key.objects.bulk_update(keys, ['name'], batch_size=1000)
        if key_shells:
            KeyShell.objects.bulk_update(key_shells, ['name'], batch_size=1000)
        if remote_shells:
            RemoteShell.objects.bulk_update(remote_shells, ['name'], batch_size=1000)
        if emergency_keys:
            EmergencyKey.objects.bulk_update(emergency_keys, ['name'], batch_size=1000)
