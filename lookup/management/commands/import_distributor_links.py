from django.core.management.base import BaseCommand
from logging import getLogger

from lookup.models import Key, KeyShell, Remote, RemoteShell, EmergencyKey, TransponderKey, Distributor, DistributorKey, DistributorTransponderKey, DistributorRemote, DistributorKeyShell, DistributorRemoteShell, DistributorEmergencyKey

import openpyxl
from datetime import datetime
import pytz

LOG = getLogger(__name__)

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str)

    def handle(self, *args, **options):
        try:
            file_name = options['file']
            if file_name:
                wb = openpyxl.load_workbook(file_name)
                sheet1 = wb[wb.sheetnames[0]]

                print('\nExcel spreadsheet loaded successfully.')

                if self._index_spreadsheet_columns(sheet1):
                    link_data = self._read_link_data_from_excel(sheet1)

                    link_dict = self._bulk_create_or_get_links(link_data)

                    timestamp = datetime.now(pytz.timezone('US/Pacific')).strftime('%Y%m%d_%H%M%S')
                    output_file = 'static/excel/output_file_' + timestamp + '.xlsx'

                    self._generate_link_data_output_workbook(link_dict, output_file)

                    print('\nImport results saved to ' + output_file + '.')
        except ValueError as e:
            pass

    def _index_spreadsheet_columns(self, sheet1):
        self.COL_INDEX_ID_1 = -1
        self.COL_INDEX_TYPE_1 = -1
        self.COL_INDEX_DISTRIBUTOR_1 = -1
        self.COL_INDEX_LINK_1 = -1

        i = 1
        while True:
            col_name = sheet1.cell(1, i).value
            if col_name == None:
                break
            elif col_name.lower() == 'id':
                self.COL_INDEX_ID_1 = i
            elif col_name.lower() == 'type':
                self.COL_INDEX_TYPE_1 = i
            elif col_name.lower() == 'distributor':
                self.COL_INDEX_DISTRIBUTOR_1 = i
            elif col_name.lower() == 'link':
                self.COL_INDEX_LINK_1 = i

            i += 1
        
        if self.COL_INDEX_ID_1 == -1:
            print('\nID column missing.')
            return False
        elif self.COL_INDEX_TYPE_1 == -1:
            print('\nType column missing.')
            return False
        elif self.COL_INDEX_DISTRIBUTOR_1 == -1:
            print('\nDistributor column missing.')
            return False
        elif self.COL_INDEX_LINK_1 == -1:
            print('\nLink column missing.')
            return False

        print('\nAll columns are present')
        return True

    def _read_link_data_from_excel(self, sheet):
        link_data = []
        for row in range(2, sheet.max_row + 1):
            if isinstance(sheet.cell(row, self.COL_INDEX_ID_1).value, float):
                id1 = int(sheet.cell(row, self.COL_INDEX_ID_1).value)
            elif isinstance(sheet.cell(row, self.COL_INDEX_ID_1).value, int):
                id1 = sheet.cell(row, self.COL_INDEX_ID_1).value
            else:
                print('ID field is invalid')
                continue

            typename = str(sheet.cell(row, self.COL_INDEX_TYPE_1).value)
            distributor = str(sheet.cell(row, self.COL_INDEX_DISTRIBUTOR_1).value)
            link = str(sheet.cell(row, self.COL_INDEX_LINK_1).value)
            link_data.append({   'id': id1,
                                    'type': typename,
                                    'distributor': distributor,
                                    'link': link })

        print('\nLink data read from excel successfully.')
        print(link_data)

        return link_data

    def _bulk_create_or_get_links(self, link_data): 
        output_links = []

        for link in link_data:
            p = self._get_or_create_link(link)
            print(p)

            output_links.append(p)

        print('\nValid links created or updated successfully.')

        return output_links

    def _get_or_create_link(self, link_data):
        if 'id' not in link_data.keys():
            raise ValueError('Missing ID field from import.')
        if 'type' not in link_data.keys():
            raise ValueError('Missing type field from import.')
        if 'distributor' not in link_data.keys():
            raise ValueError('Missing distributor field from import.')
        if 'link' not in link_data.keys():
            raise ValueError('Missing link field from import.')

        distributors = Distributor.objects.filter(name=link_data['distributor'])

        if distributors:
            distributor = distributors[0]

            if link_data.get('type').lower() == 'key' or link_data.get('type').lower() == 'transponder key':
                keys = TransponderKey.objects.filter(id=link_data.get('id'))
                if keys:
                    links = DistributorTransponderKey.objects.filter(distributor=distributor, transponder_key=keys[0])
                    if links:
                        link = links[0]
                        link.link = link_data.get('link')
                        link.save()
                        return link
                    else:
                        return DistributorTransponderKey.objects.create(distributor=distributor, transponder_key=keys[0], link=link_data.get('link'))
                else:
                    print('Transponder Key ID ' + str(link_data.get('id')) + ' not found.')
            elif link_data.get('type').lower() == 'remote':
                remotes = Remote.objects.filter(id=link_data.get('id'))
                if remotes:
                    links = DistributorRemote.objects.filter(distributor=distributor, remote=remotes[0])
                    if links:
                        link = links[0]
                        link.link = link_data.get('link')
                        link.save()
                        return link
                    else:
                        return DistributorRemote.objects.create(distributor=distributor, remote=remotes[0], link=link_data.get('link'))
                else:
                    print('Remote ID ' + str(link_data.get('id')) + ' not found.')
            elif link_data.get('type').lower() == 'key shell':
                key_shells = KeyShell.objects.filter(id=link_data.get('id'))
                if key_shells:
                    links = DistributorKeyShell.objects.filter(distributor=distributor, key_shell=key_shells[0])
                    if links:
                        link = links[0]
                        link.link = link_data.get('link')
                        link.save()
                        return link
                    else:
                        return DistributorKeyShell.objects.create(distributor=distributor, key_shell=key_shells[0], link=link_data.get('link'))
                else:
                    print('Key Shell ID ' + str(link_data.get('id')) + ' not found.')
            elif link_data.get('type').lower() == 'remote shell':
                remote_shells = RemoteShell.objects.filter(id=link_data.get('id'))
                if remote_shells:
                    links = DistributorRemoteShell.objects.filter(distributor=distributor, remote_shell=remote_shells[0])
                    if links:
                        link = links[0]
                        link.link = link_data.get('link')
                        link.save()
                        return link
                    else:
                        return DistributorRemoteShell.objects.create(distributor=distributor, remote_shell=remote_shells[0], link=link_data.get('link'))
                else:
                    print('Remote Shell ID ' + str(link_data.get('id')) + ' not found.')
            elif link_data.get('type').lower() == 'emergency key':
                emergency_keys = EmergencyKey.objects.filter(id=link_data.get('id'))
                if emergency_keys:
                    links = DistributorEmergencyKey.objects.filter(distributor=distributor, emergency_key=emergency_keys[0])
                    if links:
                        link = links[0]
                        link.link = link_data.get('link')
                        link.save()
                        return link
                    else:
                        return DistributorEmergencyKey.objects.create(distributor=distributor, emergency_key=emergency_keys[0], link=link_data.get('link'))
                else:
                    print('Emergency Key ID ' + str(link_data.get('id')) + ' not found.')
            else:
                raise ValueError('Datatype ' + link_data.get('type') + ' invalid.')
        else:
            raise ValueError('Distributor ' + link_data.get('distributor') + ' not found.')

    def _generate_link_data_output_workbook(self, links, file_name):
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']

        sheet['A1'] = 'Product ID'
        sheet['B1'] = 'Type'
        sheet['C1'] = 'Distributor'
        sheet['D1'] = 'Link'
        i = 2
        for link in links:
            if link:
                sheet['A' + str(i)] = link.id
                sheet['C' + str(i)] = link.distributor.name
                sheet['D' + str(i)] = link.link
                if isinstance(link, Remote):
                    sheet['B' + str(i)] = 'Remote'
                if isinstance(link, TransponderKey):
                    sheet['B' + str(i)] = 'Transponder Key'
                elif isinstance(link, KeyShell):
                    sheet['B' + str(i)] = 'Key Shell'
                elif isinstance(link, RemoteShell):
                    sheet['B' + str(i)] = 'Remote Shell'
                elif isinstance(link, EmergencyKey):
                    sheet['B' + str(i)] = 'Emergency Key'

                i += 1

        wb.save(file_name)
