from django.core.management.base import BaseCommand
from logging import getLogger

from lookup.models import Key, TransponderKey

import openpyxl
from datetime import datetime
import pytz

LOG = getLogger(__name__)

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str)

    def handle(self, *args, **options):
        try:
            file_name = options['file']
            if file_name:
                wb = openpyxl.load_workbook(file_name)
                sheet1 = wb[wb.sheetnames[0]]

                print('\nExcel spreadsheet loaded successfully.')

                if self._index_spreadsheet_columns(sheet1):
                    key_data = self._read_key_data_from_excel(sheet1)

                    key_dict = self._bulk_convert_to_transponder_key(key_data)

                    timestamp = datetime.now(pytz.timezone('US/Pacific')).strftime('%Y%m%d_%H%M%S')
                    output_file = 'static/excel/output_file_' + timestamp + '.xlsx'

                    self._generate_key_data_output_workbook(key_dict, output_file)

                    print('\nImport results saved to ' + output_file + '.')
        except ValueError as e:
            pass

    def _index_spreadsheet_columns(self, sheet1):
        self.COL_INDEX_ID_1 = -1
        self.COL_INDEX_SKU_1 = -1

        i = 1
        while True:
            col_name = sheet1.cell(1, i).value
            if col_name == None:
                break
            elif col_name.lower() == 'id':
                self.COL_INDEX_ID_1 = i
            elif col_name.lower() == 'sku':
                self.COL_INDEX_SKU_1 = i

            i += 1
        
        if self.COL_INDEX_ID_1 == -1:
            print('\nID column missing.')
            return False
        elif self.COL_INDEX_SKU_1 == -1:
            print('\nSKU column missing.')
            return False

        print('\nAll columns are present')
        return True

    def _read_key_data_from_excel(self, sheet):
        key_data = []
        for row in range(2, sheet.max_row + 1):
            if isinstance(sheet.cell(row, self.COL_INDEX_ID_1).value, float):
                id1 = int(sheet.cell(row, self.COL_INDEX_ID_1).value)
            elif isinstance(sheet.cell(row, self.COL_INDEX_ID_1).value, int):
                id1 = sheet.cell(row, self.COL_INDEX_ID_1).value
            else:
                print('ID field is invalid')
                continue

            sku = str(sheet.cell(row, self.COL_INDEX_SKU_1).value)
            key_data.append({   'id': id1,
                                    'sku': sku })

        print('\nKey data read from excel successfully.')
        print(key_data)

        return key_data
    
    def _bulk_convert_to_transponder_key(self, key_data):
        output_keys = []

        for key in key_data:
            TransponderKey.objects.create()
            p = self._get_or_create_link(key)
            print(p)

            output_keys.append(p)

        print('\nValid links created or updated successfully.')

        return output_keys

    def _generate_link_data_output_workbook(self, keys, file_name):
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']

        sheet['A1'] = 'ID'
        sheet['B1'] = 'SKU'
        i = 2
        for key in keys:
            if key:
                sheet['A' + str(i)] = key.id
                sheet['B' + str(i)] = key.sku

                i += 1

        wb.save(file_name)
