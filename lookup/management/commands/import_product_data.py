from django.core.management.base import BaseCommand
from logging import getLogger

from lookup.models import Key, KeyShell, Remote, RemoteShell, EmergencyKey, VehicleApplication

import openpyxl
from datetime import datetime
import pytz

LOG = getLogger(__name__)

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str)

    def handle(self, *args, **options):
        try:
            file_name = options['file']
            if file_name:
                wb = openpyxl.load_workbook(file_name)
                sheet1 = wb[wb.sheetnames[0]]
                sheet2 = wb[wb.sheetnames[1]]

                print('\nExcel spreadsheet loaded successfully.')

                if self._index_spreadsheet_columns(sheet1, sheet2):
                    product_data = self._read_product_data_from_excel(sheet1)
                    vehicle_applications = self._read_vehicle_application_from_excel(sheet2)

                    product_dict = self._bulk_create_or_get_products(product_data)
                    self._bulk_create_or_get_vehicle_applications(vehicle_applications)

                    timestamp = datetime.now(pytz.timezone('US/Pacific')).strftime('%Y%m%d_%H%M%S')
                    output_file = 'static/excel/output_file_' + timestamp + '.xlsx'

                    self._generate_product_data_output_workbook(product_dict, output_file)

                    print('\nImport results saved to ' + output_file + '.')
        except ValueError as e:
            pass

    def _index_spreadsheet_columns(self, sheet1, sheet2):
        self.COL_INDEX_SKU_1 = -1
        self.COL_INDEX_NAME_1 = -1
        self.COL_INDEX_TYPE_1 = -1

        i = 1
        while True:
            col_name = sheet1.cell(1, i).value
            if col_name == None:
                break
            elif col_name.lower() == 'sku':
                self.COL_INDEX_SKU_1 = i
            elif col_name.lower() == 'name':
                self.COL_INDEX_NAME_1 = i
            elif col_name.lower() == 'type':
                self.COL_INDEX_TYPE_1 = i

            i += 1
        
        if self.COL_INDEX_SKU_1 == -1:
            print('\nSKU column missing.')
            return False
        elif self.COL_INDEX_NAME_1 == -1:
            print('\nName column missing.')
            return False
        elif self.COL_INDEX_TYPE_1 == -1:
            print('\nType column missing.')
            return False

        self.COL_INDEX_SKU_2 = -1
        self.COL_INDEX_VEHICLE_2 = -1

        i = 1
        while True:
            col_name = sheet2.cell(1, i).value
            if col_name == None:
                break
            elif col_name.lower() == 'sku':
                self.COL_INDEX_SKU_2 = i
            elif col_name.lower() == 'vehicle':
                self.COL_INDEX_VEHICLE_2 = i
            
            i += 1

        if self.COL_INDEX_SKU_2 == -1:
            print('\nSKU column missing.')
            return False
        elif self.COL_INDEX_VEHICLE_2 == -1:
            print('\nVehicle column missing.')
            return False

        print('\nAll columns are present')
        return True

    def _read_product_data_from_excel(self, sheet):
        product_data = []
        for row in range(2, sheet.max_row + 1):
            if isinstance(sheet.cell(row, self.COL_INDEX_SKU_1).value, float):
                sku = str(int(sheet.cell(row, self.COL_INDEX_SKU_1).value))
            else:
                sku = str(sheet.cell(row, self.COL_INDEX_SKU_1).value)
            name = str(sheet.cell(row, self.COL_INDEX_NAME_1).value)
            product_type = str(sheet.cell(row, self.COL_INDEX_TYPE_1).value)
            product_data.append({   'sku': sku,
                                    'name': name,
                                    'type': product_type })

        print('\nProduct data read from excel successfully.')
        print(product_data)

        return product_data

    def _read_vehicle_application_from_excel(self, sheet):
        vehicle_applications = {}
        for row in range(2, sheet.max_row + 1):
            if isinstance(sheet.cell(row, self.COL_INDEX_SKU_2).value, float):
                sku = str(int(sheet.cell(row, self.COL_INDEX_SKU_2).value))
            else:
                sku = str(sheet.cell(row, self.COL_INDEX_SKU_2).value)
            vehicle = str(sheet.cell(row, self.COL_INDEX_VEHICLE_2).value)

            if not sku in vehicle_applications.keys():
                vehicle_applications[sku] = []

            vehicle_applications[sku].append(vehicle)
        
        print('\nVehicle application data read from excel succesfully.')
        print(vehicle_applications)

        return vehicle_applications

    def _bulk_create_or_get_products(self, product_data): 
        output_keys = []

        for product in product_data:
            p = self._get_or_create_product(product)
            print(p)

            output_keys.append(p)

        print('\nAll products created or loaded successfully.')

        return output_keys
    
    def _bulk_create_or_get_vehicle_applications(self, vehicle_dict):
        vehicle_applications = []
        for sku in vehicle_dict.keys():
            product_obj = self._get_product_by_sku(sku)
            if product_obj != None:
                for vehicle in vehicle_dict[sku]:
                    if not product_obj.vehicleapplication_set.filter(vehicle_range=vehicle):
                        if isinstance(product_obj, Key):
                            vehicle_applications.append(VehicleApplication(key_id=product_obj.id, vehicle_range=vehicle))
                        elif isinstance(product_obj, KeyShell):
                            vehicle_applications.append(VehicleApplication(key_shell_id=product_obj.id, vehicle_range=vehicle))
                        elif isinstance(product_obj, RemoteShell):
                            vehicle_applications.append(VehicleApplication(remote_shell_id=product_obj.id, vehicle_range=vehicle))
                        elif isinstance(product_obj, EmergencyKey):
                            vehicle_applications.append(VehicleApplication(emergency_key_id=product_obj.id, vehicle_range=vehicle))
        
        VehicleApplication.objects.bulk_create(vehicle_applications)

        print('\nVehicle application data created successfully.')

    def _generate_product_data_output_workbook(self, products, file_name):
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']

        sheet['A1'] = 'ID'
        sheet['B1'] = 'SKU'
        sheet['C1'] = 'Name'
        sheet['D1'] = 'Type'
        i = 2
        for product in products:
            sheet['A' + str(i)] = product.id
            sheet['B' + str(i)] = product.sku
            sheet['C' + str(i)] = product.name
            if isinstance(product, Key):
                sheet['D' + str(i)] = 'Key'
            elif isinstance(product, Remote):
                sheet['D' + str(i)] = 'Remote'
            elif isinstance(product, KeyShell):
                sheet['D' + str(i)] = 'Key Shell'
            elif isinstance(product, RemoteShell):
                sheet['D' + str(i)] = 'Remote Shell'
            elif isinstance(product, EmergencyKey):
                sheet['D' + str(i)] = 'Emergency Key'

            i += 1

        wb.save(file_name)


    def _get_or_create_product(self, product_data):
        if 'type' not in product_data.keys():
            raise ValueError('Missing type field from import.')
        if 'sku' not in product_data.keys():
            raise ValueError('Missing sku field from import.')
        if 'name' not in product_data.keys():
            raise ValueError('Missing name field from import.')

        if product_data.get('type').lower() == 'key':
            keys = Key.objects.filter(sku=product_data.get('sku'))
            if keys:
                return keys[0]
            else:
                return Key.objects.create(sku=product_data.get('sku'), name=product_data.get('name'))
        elif product_data.get('type').lower() == 'remote':
            remotes = Remote.objects.filter(sku=product_data.get('sku'))
            if remotes:
                return remotes[0]
            else:
                return Remote.objects.create(sku=product_data.get('sku'), name=product_data.get('name'))
        elif product_data.get('type').lower() == 'key shell':
            key_shells = KeyShell.objects.filter(sku=product_data.get('sku'))
            if key_shells:
                return key_shells[0]
            else:
                return KeyShell.objects.create(sku=product_data.get('sku'), name=product_data.get('name'))
        elif product_data.get('type').lower() == 'remote shell':
            remote_shells = RemoteShell.objects.filter(sku=product_data.get('sku'))
            if remote_shells:
                return remote_shells[0]
            else:
                return RemoteShell.objects.create(sku=product_data.get('sku'), name=product_data.get('name'))
        elif product_data.get('type').lower() == 'emergency key':
            emergency_keys = EmergencyKey.objects.filter(sku=product_data.get('sku'))
            if emergency_keys:
                return emergency_keys[0]
            else:
                return EmergencyKey.objects.create(sku=product_data.get('sku'), name=product_data.get('name'))
        else:
            raise ValueError('Datatype ' + product_data.get('type') + ' invalid.')

    def _get_product_by_sku(self, sku):
        remotes = Remote.objects.filter(sku=sku)
        if remotes:
            return remotes[0]
        else:
            key_shells = KeyShell.objects.filter(sku=sku)
            if key_shells:
                return key_shells[0]
            else:
                remote_shells = RemoteShell.objects.filter(sku=sku)
                if remote_shells:
                    return remote_shells[0]
                else:
                    emergency_keys = EmergencyKey.objects.filter(sku=sku)
                    if emergency_keys:
                        return emergency_keys[0]
                    else:
                        return None

    
        
