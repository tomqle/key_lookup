from django.core.management.base import BaseCommand
from logging import getLogger

from lookup.models import Key, VehicleApplication

import openpyxl

LOG = getLogger(__name__)

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str)

    def handle(self, *args, **options):
        file_name = options['file']
        if file_name:
            wb = openpyxl.load_workbook(file_name)
            sheet1 = wb['Sheet1']
            sheet2 = wb['Sheet2']

            key_names = self._read_key_name_from_excel(sheet1)
            vehicle_applications = self._read_vehicle_application_from_excel(sheet2)
            print(vehicle_applications)

            keys_dict = self._bulk_create_or_get_keys(key_names)
            self._bulk_create_or_get_vehicle_applications(vehicle_applications)

            self._generate_key_name_output_workbook(keys_dict, 'static/excel/output_file.xlsx')


    def _read_key_name_from_excel(self, sheet):
        key_names = []
        for row in range(2, sheet.max_row + 1):
            key_names.append(str(sheet['A' + str(row)].value))

        return key_names

    def _read_vehicle_application_from_excel(self, sheet):
        vehicle_applications = {}
        for row in range(2, sheet.max_row + 1):
            key_name = str(sheet['B' + str(row)].value)

            if not key_name in vehicle_applications.keys():
                vehicle_applications[key_name] = []

            vehicle_applications[key_name].append(str(sheet['A' + str(row)].value))

        return vehicle_applications


    def _bulk_create_or_get_keys(self, key_names):
        output_keys = []
        for key_name in key_names:
            keys = Key.objects.filter(name=key_name)
            key = None
            if keys:
                key = keys[0]
            else:
                key = Key.objects.create(name=key_name)

            output_keys.append({
                'id': key.id,
                'name': key.name,
            })

        return output_keys

    def _bulk_create_or_get_vehicle_applications(self, vehicles_dict):
        vehicle_applications = []
        for key in vehicles_dict.keys():
            key_obj = Key.objects.get(name=key)
            for vehicle in vehicles_dict[key]:
                if not key_obj.vehicleapplication_set.filter(vehicle_range=vehicle):
                    vehicle_applications.append(
                        VehicleApplication(key_id=key_obj.id, vehicle_range=vehicle)
                    )

        VehicleApplication.objects.bulk_create(vehicle_applications)

    def _generate_key_name_output_workbook(self, keys_dict, file_name):
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']

        sheet['A1'] = 'name'
        sheet['B1'] = 'id'
        i = 2
        for key_dict in keys_dict:
            sheet['A' + str(i)] = key_dict['name']
            sheet['B' + str(i)] = key_dict['id']
            i += 1

        wb.save(file_name)
